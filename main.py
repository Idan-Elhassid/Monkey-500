import requests
from openpyxl import load_workbook
from random import *
from tkinter import *


def choose_10_stocks():
        global stock_info
        canvas.itemconfig(canvas_image, image=card)

        wb = load_workbook("sp 500 stocks.xlsx")
        ws = wb.active

        x = randint(2, 506)
        ticker_cell = ws[f"a{x}"].value
        company_name = ws[f"b{x}"].value
        main_focus = ws[f"e{x}"].value

        stock_ticker = ticker_cell
        # your API key goes in the next line...
        api_key = ""
        stock_endpoint = "https://www.alphavantage.co/query/"

        stock_params = {
                "function": "TIME_SERIES_DAILY",
                "symbol": stock_ticker,
                "apikey": api_key,

        }
        response = requests.get(stock_endpoint, params=stock_params)
        data = response.json()['Time Series (Daily)']
        data_list = [value for (key, value) in data.items()]
        yesterday_data = data_list[0]
        yesterday_closing_price = yesterday_data["4. close"]

        canvas.itemconfig(stock_info, text=f"WE Choose A Stock for you!\nits Ticker symbol is: {stock_ticker}\nthe company name is: {company_name}\nits a company in the field of: {main_focus}\n\nits stock price as of yesterday is: ${yesterday_closing_price}", font=("david", 18, "bold"))


window = Tk()
window.title("monkey 500")
window.config(padx=100, pady=50)

canvas = Canvas(width=700, height=500, highlightthickness=0)
monkey = PhotoImage(file="twomon.PNG")
card = PhotoImage(file="card.png")
go_button = PhotoImage(file="letts go.png")
canvas_image = canvas.create_image(350, 250, image=monkey)
stock_info = canvas.create_text(350, 250, text="")
canvas.grid(column=0, row=1)

welcome_label = Label(text="Hello, and welcome to: MONKEY 500\n The monkeys will choose GREAT Stocks for you", font=("david", 40))
welcome_label.grid(column=0, row=0, columnspan=2)

choose_stock_button = Button(image=go_button, bd=0, command=choose_10_stocks)
choose_stock_button.grid(column=1, row=1)


window.mainloop()

